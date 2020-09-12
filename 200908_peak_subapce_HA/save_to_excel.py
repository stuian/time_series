import openpyxl
import os

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    if os.path.exists(path):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.create_sheet(sheet_name)
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
        workbook.save(path)
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
        workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


def main():
    book_name_xlsx = 'xlsx格式测试工作簿.xlsx'

    sheet_name_xlsx = 'xlsx格式测试表'

    # value3 = [["姓名", "性别", "年龄", "城市", "职业"],
    #           ["111", "女", "66", "石家庄", "运维工程师"],
    #           ["222", "男", "55", "南京", "饭店老板"],
    #           ["333", "女", "27", "苏州", "保安"], ]
    value3 = [['id','分数'],[1,65]]

    write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
    read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)

if __name__ == '__main__':
    main()